import time
import pyexcel
import smtplib
import datetime
import openpyxl
import schedule
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email.mime.multipart import MIMEMultipart

def MailSender(C_name,C_GSTIN,C_Inv_no,C_Inv_val,time) :
    try :
        fromaddr = ""
        toaddr = ""

        msg = MIMEMultipart()

        msg['From'] = fromaddr
        msg['To'] = toaddr

        body = """
        Hello %s,
        It's a reminder for you.
        Today %s you have to make a bill of customer .
        Details are given below :

        Name of Customer : %s
        GSTIN of Customer : %s
        Invoice No : %s 
        Invoice Value : %s



        Thanks & Regards,
        Rutuja Jamdar
       

        """%(toaddr,datetime.date.today(),C_name,C_GSTIN,C_Inv_no,C_Inv_val)

        subject = """
        Reminder for a bill scheduled on %s

        """%(datetime.date.today())

        msg['Subject'] = subject

        msg.attach(MIMEText(body,'plain'))

        s = smtplib.SMTP('smtp.gmail.com',587)

        s.starttls()

        s.login(fromaddr,"")

        text = msg.as_string()

        s.sendmail(fromaddr,toaddr,text)

        s.quit()

        print("Reminder send succesfully")

        
    
    except Exception as E :
        print("Unable to send mail",E)


def Reminder_Application() :
    date = datetime.date.today()

    file_name = "demo.xlsx"

    wb = openpyxl.load_workbook(file_name)
    ws = wb.active

    for i in range(2, 8):
        flag = False
        info = []
        for j in range(1, 6):

            cell_obj = ws.cell(row=i, column=j)
            d=str((cell_obj.value)).split(" ")[0]

            if(j==1 and str(date)==d) :
                flag = True
                    
            elif flag == True :
                    cell_obj = ws.cell(row=i, column=j)
                    info.append(cell_obj.value)

        if flag == True :
            MailSender(info[0],info[2],info[1],info[3],datetime.time) 
            print("Mail send successfully")   
                
def main() :
    schedule.every().day.at_time("07:00").do(Reminder_Application)
    while True :
        schedule.run_pending()
        time.sleep(1)

if __name__ == "__main__" :
    main()
        
